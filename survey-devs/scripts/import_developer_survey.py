#!/usr/bin/env python3
"""Import Developer Survey responses from a Microsoft Forms Excel export.

Reads: respostas-survey-devs.xlsx (or the path passed via --input)
Writes:
  - survey-devs/respostas-devs.json (schema of survey-devs/respostas-mock-devs.json)
  - saida/import-survey-log-<DATE>.md (import log)

Rules (deterministic, mirrors /import-developer-survey skill):
  - Columns matched by header regex S<n>-Q<n> (column-order agnostic).
  - The survey is ANONYMOUS: names/emails are NEVER copied to the output.
    If the Email/Name columns contain data the script warns (Forms anonymity
    was not enforced) but still discards them.
  - Question types (choice / multi / text) come from the question bank
    survey-devs/perguntas-para-forms-devs.md (75 questions in 9 sections).
  - Raw answer text is stored as-is; multi answers stay semicolon-joined
    (downstream analysis splits them).

Usage:
    python3 import_developer_survey.py
    python3 import_developer_survey.py --input caminho/para/respostas-survey-devs.xlsx
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
KIT = SCRIPT_DIR.parent.parent  # kit root

QID_PATTERN = re.compile(r"^\s*(S[1-9]-Q\d+)\b")
BANK_HEADING = re.compile(r"###\s+Pergunta\s+`(S\d+-Q\d+)`\s+[—–-]+\s+_(.+?)_")
NAME_HEADERS = {"name", "nome", "nombre"}
EMAIL_HEADERS = {"email", "e-mail", "correo", "correo electronico", "correo electrónico"}


def fail(msg: str) -> "sys.NoReturn":
    print(f"❌ {msg}")
    sys.exit(1)


def load_workbook_safe(path: Path):
    try:
        import openpyxl
    except ImportError:
        fail("openpyxl não está instalado. Rode: pip install openpyxl")
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        fail(f"Não consegui abrir '{path}' como .xlsx ({exc.__class__.__name__}). "
             f"Verifique se o arquivo é um export válido do Microsoft Forms.")


def question_types() -> dict:
    """qid -> choice | multi | text, parsed from the canonical question bank."""
    bank = KIT / "survey-devs" / "perguntas-para-forms-devs.md"
    if not bank.exists():
        fail(f"Banco de questões não encontrado: {bank}")
    mapping = {}
    for qid, kind in BANK_HEADING.findall(bank.read_text(encoding="utf-8")):
        low = kind.lower()
        if "multiple" in low:
            mapping[qid] = "multi"
        elif "single" in low:
            mapping[qid] = "choice"
        elif low.startswith("short text"):
            mapping[qid] = "text-short"
        else:
            mapping[qid] = "text"
    if not mapping:
        fail(f"Não consegui extrair os tipos de questão de {bank}.")
    return mapping


def map_columns(header_row):
    col_to_qid = {}
    name_col = email_col = None
    for idx, cell in enumerate(header_row, start=1):
        val = str(cell.value or "").strip()
        if not val:
            continue
        m = QID_PATTERN.match(val)
        if m:
            col_to_qid[idx] = m.group(1)
            continue
        low = val.lower()
        if low in NAME_HEADERS and name_col is None:
            name_col = idx
        elif low in EMAIL_HEADERS and email_col is None:
            email_col = idx
    return col_to_qid, name_col, email_col


def main():
    ap = argparse.ArgumentParser(
        description="Importa o Developer Survey (export .xlsx do Microsoft Forms, ANÔNIMO) "
                    "para survey-devs/respostas-devs.json.")
    ap.add_argument("--input", default=str(KIT / "respostas-survey-devs.xlsx"),
                    help="Caminho do .xlsx (default: respostas-survey-devs.xlsx na raiz)")
    ap.add_argument("--output", default=str(KIT / "survey-devs" / "respostas-devs.json"),
                    help="JSON de saída (default: survey-devs/respostas-devs.json)")
    ap.add_argument("--log-dir", default=str(KIT / "saida"),
                    help="Diretório do import-log (default: saida/)")
    args = ap.parse_args()

    xlsx = Path(args.input)
    out_path = Path(args.output)
    log_dir = Path(args.log_dir)

    if not xlsx.exists():
        fail(f"Não encontrei o arquivo '{xlsx}'. Veja survey-devs/INSTRUCOES-FORMS-DEVS.md "
             f"ou rode com --input <caminho-do-xlsx>.")

    types = question_types()
    expected = len(types)  # 75 questions in the canonical bank

    wb = load_workbook_safe(xlsx)
    ws = wb.active
    col_to_qid, name_col, email_col = map_columns(ws[1])

    if not col_to_qid:
        fail("Nenhum header no padrão S<n>-Q<n> encontrado na linha 1. "
             "Confirme que o arquivo é o export do Developer Survey.")

    warnings = []
    if len(col_to_qid) != expected:
        warnings.append(f"{len(col_to_qid)} questões detectadas no Excel (esperado {expected} "
                        f"do banco canônico): verifique o formato dos headers")
    unknown = sorted(set(col_to_qid.values()) - set(types))
    if unknown:
        warnings.append(f"questões fora do banco canônico: {', '.join(unknown)}, "
                        f"importadas com type 'text'")

    respondents = []
    pii_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        def cell_value(col):
            return row[col - 1].value if col and col <= len(row) else None

        # Anonymity check: warn if the Forms export leaked identity, then DISCARD it.
        # Anonymous Forms exports fill Email/Name with the literal "anonymous".
        def is_identity(col):
            s = str(cell_value(col) or "").strip().lower()
            return s not in ("", "anonymous", "anonymous user", "anônimo", "anonimo")

        if is_identity(name_col) or is_identity(email_col):
            pii_rows += 1

        responses = {}
        for col, qid in col_to_qid.items():
            raw = cell_value(col)
            if raw is None or str(raw).strip() == "":
                continue
            responses[qid] = {"type": types.get(qid, "text"), "value": str(raw).strip()}
        if not responses:
            continue  # empty row
        respondents.append({"respondent_id": len(respondents) + 1, "responses": responses})

    if not respondents:
        fail("Nenhuma linha de respondente encontrada (linhas 2+ estão vazias).")
    if pii_rows:
        warnings.append(f"{pii_rows} linha(s) com Email/Name preenchidos: o Forms não estava "
                        f"em modo anônimo; os dados de identidade foram DESCARTADOS no import")

    output = {
        "metadata": {
            "source": xlsx.name,
            "imported_at": datetime.datetime.now(datetime.timezone.utc)
                           .isoformat(timespec="seconds").replace("+00:00", "Z"),
            "total_respondents": len(respondents),
            "total_questions": expected,
            "total_questions_detected": len(col_to_qid),
            "anonymous": True,
        },
        "respondents": respondents,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n",
                        encoding="utf-8")

    # Per-section coverage for the log
    sections = sorted({q.split("-")[0] for q in col_to_qid.values()})
    sec_lines = []
    for sec in sections:
        sec_qids = [q for q in col_to_qid.values() if q.startswith(sec + "-")]
        possible = len(sec_qids) * len(respondents)
        got = sum(1 for r in respondents for q in sec_qids if q in r["responses"])
        pct = round(100 * got / possible) if possible else 0
        sec_lines.append(f"| {sec} | {got}/{possible} | {pct}% |")

    today = datetime.date.today().isoformat()
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"import-survey-log-{today}.md"
    log_lines = [
        f"# Import log - Developer Survey ({today})",
        "",
        "## Resumo",
        f"- Arquivo: {xlsx.name}",
        f"- Respondentes: {len(respondents)} (anônimos)",
        f"- Questões detectadas: {len(col_to_qid)} / {expected}",
        f"- Saída: {out_path.relative_to(KIT) if out_path.is_relative_to(KIT) else out_path}",
        "",
        "## Cobertura por seção",
        "| Seção | Respostas | % |",
        "|---|---|---|",
        *sec_lines,
        "",
        "## Alertas",
        *([f"- {w}" for w in warnings] or ["- (nenhum)"]),
        "",
        "## Próximo",
        "Rode `/insights-developer-survey` para gerar relatório agregado.",
        "",
    ]
    log_path.write_text("\n".join(log_lines), encoding="utf-8")

    total_cells = len(col_to_qid) * len(respondents)
    got_cells = sum(len(r["responses"]) for r in respondents)
    coverage = round(100 * got_cells / total_cells) if total_cells else 0
    print(f"✓ Survey importado → {out_path}")
    print(f"   {len(respondents)} respondentes anônimos, {len(col_to_qid)} questões processadas")
    print(f"✓ Log: {log_path}")
    print(f"   Cobertura: {coverage}%")
    if warnings:
        print(f"⚠ {len(warnings)} alerta(s), ver log")
    print("Próximo: /insights-developer-survey")
    return 0


if __name__ == "__main__":
    sys.exit(main())
