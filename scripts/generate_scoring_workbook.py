#!/usr/bin/env python3
"""Generate the full auditable scoring workbook (pontuacao-e-calculo.xlsx).

Builds an Excel workbook covering ALL questions and capabilities in
framework.json (the single source of truth for weights), with live
SUMPRODUCT formulas that mirror referencia/pontuacao-e-calculo.md 1:1:

  1. capability score = SUMPRODUCT(levels, weights) / sum of answered weights
  2. pillar score     = SUMPRODUCT(cap scores, cap weights) over scored caps
  3. overall score    = same SUMPRODUCT over ALL capabilities (NOT pillar mean)

Sheets:
  - "Leia-me"   what the workbook is, how scores compute, label/threshold tables
  - "Respostas" one row per question: qid, pillar, capability, question text
                (PT-BR, from referencia/P*.md), weight (from framework.json),
                level input cell, evidence cell
  - "Cálculo"   per-capability SUMPRODUCT rows, pillar rows, overall row,
                maturity label lookup and coverage threshold counters

Deterministic: the same inputs produce byte-identical output (fixed document
properties, fixed zip timestamps, stable ordering, no run timestamps in cells).

Usage:
    python3 scripts/generate_scoring_workbook.py                       # empty template
    python3 scripts/generate_scoring_workbook.py --respostas respostas.json \
        --out saida/pontuacao-preenchida-2026-05-08.xlsx               # populated
    python3 scripts/generate_scoring_workbook.py --self-check          # verify formulas
"""
from __future__ import annotations

import argparse
import datetime
import io
import json
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is not installed: run 'make install-deps' "
          "(or: pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
KIT_ROOT = SCRIPT_DIR.parent

# Maturity labels (platform data labels; must match scripts/compute_scores.py verbatim).
LABELS = ["L0 — Inicial", "L1 — Em Desenvolvimento", "L2 — Definido",
          "L3 — Gerenciado", "L4 — Otimizando"]
LABEL_NO_ANSWER = "Sem resposta"
LABEL_CUTS = [0, 0.5, 1.5, 2.5, 3.5]

# Fixed timestamp for document properties so regeneration is byte-stable.
FIXED_STAMP = datetime.datetime(2026, 1, 1, 0, 0, 0)

QUESTION_HEADING = re.compile(r"^###\s+(P\d+-C\d+-Q\d+)\s+—\s+(.+?)\s*$")

# Styles ---------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFFF2CC")
SECTION_FONT = Font(bold=True, size=12)
THIN = Side(style="thin", color="FFB0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def fail(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def warn(msg: str) -> None:
    print(f"WARNING: {msg}", file=sys.stderr)


def rel_to_kit(path: Path) -> str:
    try:
        return path.resolve().relative_to(KIT_ROOT).as_posix()
    except ValueError:
        return str(path)


def load_json(path: Path, hint: str) -> dict:
    if not path.exists():
        fail(f"{rel_to_kit(path)} missing: {hint}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"{rel_to_kit(path)} is not valid JSON (line {exc.lineno}): fix the syntax and retry")
    return {}  # unreachable


def load_question_texts(framework: dict) -> dict[str, str]:
    """Parse '### <QID> — <text>' headings from referencia/P*.md.

    referencia/P1-*.md, P2-*.md and P3-*.md hold the official PT-BR question
    wording (framework.json only carries ids/weights/flags)."""
    texts: dict[str, str] = {}
    doc_paths = sorted((KIT_ROOT / "referencia").glob("P[123]-*.md"))
    if len(doc_paths) != 3:
        fail("expected 3 pillar question docs at referencia/P1-*.md, P2-*.md, P3-*.md; "
             f"found {len(doc_paths)}")
    for doc in doc_paths:
        for line in doc.read_text(encoding="utf-8").splitlines():
            match = QUESTION_HEADING.match(line)
            if match:
                texts[match.group(1)] = match.group(2)
    expected = [q["id"] for p in framework["pillars"]
                for c in p["capabilities"] for q in c["questions"]]
    missing = [qid for qid in expected if qid not in texts]
    if missing:
        fail(f"{len(missing)} question(s) have no '### QID — text' heading in "
             f"referencia/P*.md: {', '.join(missing[:5])}"
             f"{'...' if len(missing) > 5 else ''}")
    return texts


def load_responses(path: Path | None) -> dict[str, dict]:
    if path is None:
        return {}
    respostas = load_json(path, "point --respostas at a filled respostas.json")
    responses = respostas.get("responses")
    if not isinstance(responses, dict):
        fail(f"{rel_to_kit(path)} has no 'responses' object: see respostas.json.example")
    for qid, entry in responses.items():
        if entry is None:
            continue
        if not isinstance(entry, dict):
            fail(f"{rel_to_kit(path)}: question {qid} must be an object like "
                 f'{{"level": 2, "evidence": "..."}}, got {entry!r}')
        level = entry.get("level")
        if level is None:
            continue
        if isinstance(level, bool) or not isinstance(level, (int, float)) or not 0 <= level <= 4:
            fail(f"{rel_to_kit(path)}: question {qid} has level {level!r}: "
                 "use null or a number in [0, 4]")
    return responses


def label_lookup_formula(score_ref: str) -> str:
    """Excel LOOKUP over the 0.5/1.5/2.5/3.5 cut points."""
    cuts = "{" + ",".join(str(c) for c in LABEL_CUTS) + "}"
    labels = "{" + ",".join(f'"{label}"' for label in LABELS) + "}"
    return f"LOOKUP({score_ref},{cuts},{labels})"


def label_for(score: float) -> str:
    """Python mirror of the LOOKUP formula (and of compute_scores.label_for)."""
    if score < 0.5:
        return LABELS[0]
    if score < 1.5:
        return LABELS[1]
    if score < 2.5:
        return LABELS[2]
    if score < 3.5:
        return LABELS[3]
    return LABELS[4]


# Workbook builders ----------------------------------------------------------

def build_leiame(wb: Workbook, total_questions: int, total_caps: int) -> None:
    ws = wb.create_sheet("Leia-me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    def put(row: int, col: int, value, font: Font | None = None) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        cell.alignment = WRAP

    put(1, 1, "Planilha auditável de pontuação (AI Maturity Assessment)",
        Font(bold=True, size=14))
    put(3, 1, "O que é", SECTION_FONT)
    put(4, 1, "", None)
    put(4, 2, f"Esta planilha cobre todas as {total_questions} questões e "
              f"{total_caps} capabilities do framework.json e calcula os scores com "
              "fórmulas SUMPRODUCT visíveis, idênticas ao algoritmo oficial descrito em "
              "referencia/pontuacao-e-calculo.md e implementado em scripts/compute_scores.py.")
    put(6, 1, "Como usar", SECTION_FONT)
    put(7, 2, '1. Na aba "Respostas", preencha a coluna "Nível" (0 a 4) para cada '
              "questão respondida. Deixe em branco as questões sem resposta.")
    put(8, 2, '2. Registre a evidência de cada resposta na coluna "Evidência".')
    put(9, 2, '3. A aba "Cálculo" atualiza automaticamente: score por capability, '
              "por pilar, overall, rótulos de maturidade e o threshold de cobertura.")
    put(10, 2, "4. Não edite as colunas de peso: os pesos vêm do framework.json, "
               "a fonte única da verdade. Para regenerar ou pré-popular esta planilha use "
               "scripts/generate_scoring_workbook.py.")
    put(12, 1, "Como o score é calculado", SECTION_FONT)
    put(13, 2, "Capability: soma(nível x peso da questão) / soma(pesos das questões "
               "respondidas). Questões em branco não somam no numerador nem no denominador.")
    put(14, 2, "Pilar: soma(score da capability x peso da capability) / soma(pesos), "
               "considerando apenas capabilities com ao menos 1 resposta.")
    put(15, 2, "Overall: mesmo SUMPRODUCT aplicado sobre TODAS as capabilities do "
               "framework. Não é a média dos 3 pilares.")

    row = 17
    put(row, 1, "Rótulos de maturidade", SECTION_FONT)
    row += 1
    ranges = ["score < 0.5", "0.5 <= score < 1.5", "1.5 <= score < 2.5",
              "2.5 <= score < 3.5", "score >= 3.5"]
    ws.cell(row=row, column=1, value="Faixa de score").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Rótulo").font = Font(bold=True)
    for rng, label in zip(ranges, LABELS):
        row += 1
        ws.cell(row=row, column=1, value=rng).border = BOX
        ws.cell(row=row, column=2, value=label).border = BOX

    row += 2
    put(row, 1, "Threshold de cobertura", SECTION_FONT)
    row += 1
    ws.cell(row=row, column=1, value="Questões respondidas").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Status").font = Font(bold=True)
    thresholds = [(">= 40", "OK: scoring normal"),
                  ("25 a 39", "WARNING: resultado preliminar, confiabilidade limitada"),
                  ("< 25", "BLOCKED: cobertura insuficiente, não usar para decisões executivas")]
    for rng, status in thresholds:
        row += 1
        ws.cell(row=row, column=1, value=rng).border = BOX
        ws.cell(row=row, column=2, value=status).border = BOX


def build_respostas(wb: Workbook, framework: dict, texts: dict[str, str],
                    responses: dict[str, dict]) -> dict[str, tuple[int, int]]:
    """Create the Respostas sheet. Returns {capability_id: (first_row, last_row)}."""
    ws = wb.create_sheet("Respostas")
    headers = ["QID", "Pilar", "Capability", "Pergunta", "Peso", "Nível", "Evidência"]
    widths = [12, 7, 10, 70, 7, 8, 60]
    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    validation = DataValidation(
        type="decimal", operator="between", formula1="0", formula2="4",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Nível inválido",
        error="Use um número entre 0 e 4 (frações são permitidas para médias "
              "multi-respondente) ou deixe em branco.")
    ws.add_data_validation(validation)

    cap_rows: dict[str, tuple[int, int]] = {}
    row = 2
    for pillar in framework["pillars"]:
        for cap in pillar["capabilities"]:
            first = row
            for question in cap["questions"]:
                qid = question["id"]
                entry = responses.get(qid) or {}
                ws.cell(row=row, column=1, value=qid).border = BOX
                ws.cell(row=row, column=2, value=pillar["id"]).border = BOX
                ws.cell(row=row, column=3, value=cap["id"]).border = BOX
                cell = ws.cell(row=row, column=4, value=texts[qid])
                cell.border = BOX
                cell.alignment = WRAP
                ws.cell(row=row, column=5, value=question.get("weight", 1.0)).border = BOX
                level_cell = ws.cell(row=row, column=6)
                if entry.get("level") is not None:
                    level_cell.value = entry["level"]
                level_cell.fill = INPUT_FILL
                level_cell.border = BOX
                validation.add(level_cell)
                evidence_cell = ws.cell(row=row, column=7)
                if entry.get("evidence"):
                    evidence_cell.value = str(entry["evidence"])
                evidence_cell.fill = INPUT_FILL
                evidence_cell.border = BOX
                evidence_cell.alignment = WRAP
                row += 1
            cap_rows[cap["id"]] = (first, row - 1)
    return cap_rows


def build_calculo(wb: Workbook, framework: dict,
                  cap_rows: dict[str, tuple[int, int]]) -> dict:
    """Create the Cálculo sheet. Returns the cell layout used by --self-check."""
    ws = wb.create_sheet("Cálculo")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    for col in "CDEFGHIJKL":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["J"].width = 24
    ws.freeze_panes = "A4"

    ws.cell(row=1, column=1, value="Cálculo oficial (SUMPRODUCT sobre a aba Respostas)"
            ).font = SECTION_FONT

    cap_headers = ["ID", "Nome", "Pilar", "Peso", "Respondidas", "Aplicáveis",
                   "Soma ponderada", "Peso respondido", "Score", "Rótulo",
                   "Contrib. numerador", "Contrib. denominador"]
    for col, header in enumerate(cap_headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    layout: dict = {"cap_rows": {}, "pillar_rows": {}}
    row = 4
    pillar_cap_span: dict[str, tuple[int, int]] = {}
    for pillar in framework["pillars"]:
        span_first = row
        for cap in pillar["capabilities"]:
            first, last = cap_rows[cap["id"]]
            levels = f"Respostas!F{first}:F{last}"
            weights = f"Respostas!E{first}:E{last}"
            ws.cell(row=row, column=1, value=cap["id"])
            ws.cell(row=row, column=2, value=cap.get("name_pt_br", cap["id"]))
            ws.cell(row=row, column=3, value=pillar["id"])
            ws.cell(row=row, column=4, value=cap.get("weight", 1.0))
            ws.cell(row=row, column=5, value=f"=COUNT({levels})")
            ws.cell(row=row, column=6, value=last - first + 1)
            ws.cell(row=row, column=7, value=f"=SUMPRODUCT({levels},{weights})")
            ws.cell(row=row, column=8, value=f'=SUMPRODUCT(--({levels}<>""),{weights})')
            score = ws.cell(row=row, column=9, value=f'=IF(H{row}=0,"",G{row}/H{row})')
            score.number_format = "0.000"
            ws.cell(row=row, column=10,
                    value=f'=IF(I{row}="","{LABEL_NO_ANSWER}",'
                          f"{label_lookup_formula(f'I{row}')})")
            ws.cell(row=row, column=11, value=f"=IF(H{row}=0,0,D{row}*I{row})")
            ws.cell(row=row, column=12, value=f"=IF(H{row}=0,0,D{row})")
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = BOX
            layout["cap_rows"][cap["id"]] = row
            row += 1
        pillar_cap_span[pillar["id"]] = (span_first, row - 1)
    cap_first, cap_last = 4, row - 1

    row += 2
    ws.cell(row=row, column=1, value="Score por pilar").font = SECTION_FONT
    row += 1
    for col, header in enumerate(["ID", "Nome", "Score", "Rótulo"], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1
    for pillar in framework["pillars"]:
        first, last = pillar_cap_span[pillar["id"]]
        num, den = f"SUM(K{first}:K{last})", f"SUM(L{first}:L{last})"
        ws.cell(row=row, column=1, value=pillar["id"])
        ws.cell(row=row, column=2, value=pillar.get("name_pt_br", pillar["id"]))
        score = ws.cell(row=row, column=3, value=f"=IF({den}=0,0,{num}/{den})")
        score.number_format = "0.000"
        ws.cell(row=row, column=4, value=f"={label_lookup_formula(f'C{row}')}")
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BOX
        layout["pillar_rows"][pillar["id"]] = row
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Overall (SUMPRODUCT sobre TODAS as capabilities, "
                                     "não é média dos pilares)").font = SECTION_FONT
    row += 1
    num, den = f"SUM(K{cap_first}:K{cap_last})", f"SUM(L{cap_first}:L{cap_last})"
    ws.cell(row=row, column=1, value="Overall").font = Font(bold=True)
    score = ws.cell(row=row, column=3, value=f"=IF({den}=0,0,{num}/{den})")
    score.number_format = "0.000"
    ws.cell(row=row, column=4, value=f"={label_lookup_formula(f'C{row}')}")
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = BOX
    layout["overall_row"] = row

    row += 2
    ws.cell(row=row, column=1, value="Threshold de cobertura").font = SECTION_FONT
    row += 1
    total_questions = sum(last - first + 1 for first, last in cap_rows.values())
    last_resp_row = 1 + total_questions
    ws.cell(row=row, column=1, value="Respondidas")
    ws.cell(row=row, column=3, value=f"=COUNT(Respostas!F2:F{last_resp_row})")
    layout["answered_row"] = row
    row += 1
    ws.cell(row=row, column=1, value="Aplicáveis")
    ws.cell(row=row, column=3, value=total_questions)
    row += 1
    ws.cell(row=row, column=1, value="Status")
    answered_ref = f"C{layout['answered_row']}"
    ws.cell(row=row, column=3,
            value=f'=IF({answered_ref}>=40,"OK",IF({answered_ref}>=25,"WARNING","BLOCKED"))')
    layout["status_row"] = row
    return layout


def build_workbook(framework: dict, responses: dict[str, dict]) -> tuple[Workbook, dict]:
    texts = load_question_texts(framework)
    known_qids = set(texts)
    unknown = sorted(qid for qid in responses if qid not in known_qids)
    if unknown:
        warn(f"respostas contains {len(unknown)} question id(s) not in framework.json "
             f"(ignored): {', '.join(unknown[:5])}{'...' if len(unknown) > 5 else ''}")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    total_questions = sum(len(c["questions"])
                          for p in framework["pillars"] for c in p["capabilities"])
    total_caps = sum(len(p["capabilities"]) for p in framework["pillars"])
    build_leiame(wb, total_questions, total_caps)
    cap_rows = build_respostas(wb, framework, texts, responses)
    layout = build_calculo(wb, framework, cap_rows)
    layout["cap_question_rows"] = cap_rows

    wb.properties.created = FIXED_STAMP
    wb.properties.modified = FIXED_STAMP
    wb.properties.creator = "scripts/generate_scoring_workbook.py"
    wb.properties.lastModifiedBy = "scripts/generate_scoring_workbook.py"
    return wb, layout


def save_deterministic(wb: Workbook, path: Path) -> None:
    """Save the workbook, then repack the zip with fixed entry timestamps and a
    fixed dcterms timestamp so two runs over the same inputs are byte-identical.

    openpyxl overwrites dcterms:created/modified with the wall clock at save
    time (ignoring wb.properties), so docProps/core.xml is patched here."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = FIXED_STAMP.strftime("%Y-%m-%dT%H:%M:%SZ")
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(buffer) as zin, \
            zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = re.sub(r'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:)',
                              rf"\g<1>{stamp}\g<2>", text)
                data = text.encode("utf-8")
            stable = zipfile.ZipInfo(info.filename, date_time=(1980, 1, 1, 0, 0, 0))
            stable.compress_type = zipfile.ZIP_DEFLATED
            stable.external_attr = 0o644 << 16
            zout.writestr(stable, data)


# Self-check -----------------------------------------------------------------

def rederive_from_workbook(path: Path, framework: dict, layout: dict) -> dict:
    """Re-derive every SUMPRODUCT in Python from the workbook's own cell inputs
    (levels and weights on the Respostas sheet), replicating the formulas."""
    wb = load_workbook(path)
    ws = wb["Respostas"]
    caps = []
    for pillar in framework["pillars"]:
        for cap in pillar["capabilities"]:
            first, last = layout["cap_question_rows"][cap["id"]]
            num = den = 0.0
            answered = 0
            for row in range(first, last + 1):
                level = ws.cell(row=row, column=6).value
                weight = ws.cell(row=row, column=5).value
                if level is None:
                    continue
                answered += 1
                num += float(level) * float(weight)
                den += float(weight)
            score = None if den == 0 else num / den
            caps.append({"id": cap["id"], "pillar_id": pillar["id"],
                         "weight": cap.get("weight", 1.0),
                         "score": score, "answered": answered})
    pillars = {}
    for pillar in framework["pillars"]:
        scored = [(c["score"], c["weight"]) for c in caps
                  if c["pillar_id"] == pillar["id"] and c["score"] is not None]
        num = sum(s * w for s, w in scored)
        den = sum(w for _, w in scored)
        pillars[pillar["id"]] = 0.0 if den == 0 else num / den
    scored_all = [(c["score"], c["weight"]) for c in caps if c["score"] is not None]
    den = sum(w for _, w in scored_all)
    overall = 0.0 if den == 0 else sum(s * w for s, w in scored_all) / den
    return {"capabilities": caps, "pillars": pillars, "overall": overall,
            "answered": sum(c["answered"] for c in caps)}


def run_self_check(framework: dict, respostas_path: Path) -> int:
    """Generate a populated workbook, re-derive its formulas in Python, and
    assert equality with scripts/compute_scores.py on the same input."""
    responses = load_responses(respostas_path)
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        xlsx_path = tmp_dir / "workbook.xlsx"
        wb, layout = build_workbook(framework, responses)
        save_deterministic(wb, xlsx_path)

        scores_path = tmp_dir / "scores.json"
        result = subprocess.run(
            [sys.executable, str(SCRIPT_DIR / "compute_scores.py"),
             "--respostas", str(respostas_path), "--out", str(scores_path),
             "--now", "2026-01-01T00:00:00Z"],
            capture_output=True, text=True)
        if result.returncode != 0:
            fail(f"compute_scores.py failed during self-check:\n{result.stderr.strip()}")
        expected = json.loads(scores_path.read_text(encoding="utf-8"))
        derived = rederive_from_workbook(xlsx_path, framework, layout)

    errors: list[str] = []

    def check(name: str, got, want) -> None:
        if got != want:
            errors.append(f"{name}: workbook formula gives {got!r}, compute_scores gives {want!r}")

    check("overall.score", round(derived["overall"], 3), expected["overall"]["score"])
    check("overall.label", label_for(derived["overall"]), expected["overall"]["label"])
    check("threshold.answered", derived["answered"], expected["threshold"]["answered"])
    for pillar in expected["pillars"]:
        got = derived["pillars"][pillar["id"]]
        check(f"pillar {pillar['id']}.score", round(got, 3), pillar["score"])
        check(f"pillar {pillar['id']}.label", label_for(got), pillar["label"])
    derived_caps = {c["id"]: c for c in derived["capabilities"]}
    for cap in expected["capabilities"]:
        got = derived_caps[cap["id"]]
        got_score = None if got["score"] is None else round(got["score"], 3)
        check(f"capability {cap['id']}.score", got_score, cap["score"])
        got_label = LABEL_NO_ANSWER if got["score"] is None else label_for(got["score"])
        check(f"capability {cap['id']}.label", got_label, cap["label"])
        check(f"capability {cap['id']}.answered", got["answered"], cap["answered"])

    if errors:
        for error in errors:
            print(f"SELF-CHECK FAIL: {error}", file=sys.stderr)
        return 1
    n_caps = len(expected["capabilities"])
    print(f"Self-check PASSED: workbook formulas match compute_scores.py exactly "
          f"(overall={expected['overall']['score']}, "
          f"{len(expected['pillars'])} pillars, {n_caps} capabilities, "
          f"answered={expected['threshold']['answered']}) "
          f"using {rel_to_kit(respostas_path)}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate the full auditable scoring workbook from framework.json "
                    "(all questions, live SUMPRODUCT formulas).")
    parser.add_argument("--out", default=str(KIT_ROOT / "referencia" / "pontuacao-e-calculo.xlsx"),
                        help="Output .xlsx path (default: referencia/pontuacao-e-calculo.xlsx)")
    parser.add_argument("--framework", default=str(KIT_ROOT / "framework.json"),
                        help="Path to framework.json (default: kit root framework.json)")
    parser.add_argument("--respostas", default=None,
                        help="Optional respostas.json to pre-populate levels and evidence "
                             "(default: empty template)")
    parser.add_argument("--self-check", action="store_true",
                        help="Verify formula correctness: build a populated workbook in a "
                             "temp dir, re-derive every SUMPRODUCT in Python and assert "
                             "equality with scripts/compute_scores.py (uses --respostas, "
                             "or respostas.json.example when omitted); does not write --out")
    args = parser.parse_args()

    framework = load_json(Path(args.framework), "the kit framework file was not found")
    if not isinstance(framework.get("pillars"), list) or not framework["pillars"]:
        fail("framework.json has no 'pillars' list: the file is corrupted or is not the kit framework")

    if args.self_check:
        respostas_path = Path(args.respostas) if args.respostas \
            else KIT_ROOT / "respostas.json.example"
        return run_self_check(framework, respostas_path)

    responses = load_responses(Path(args.respostas) if args.respostas else None)
    wb, _ = build_workbook(framework, responses)
    out_path = Path(args.out)
    save_deterministic(wb, out_path)

    total_questions = sum(len(c["questions"])
                          for p in framework["pillars"] for c in p["capabilities"])
    answered = sum(1 for entry in responses.values()
                   if isinstance(entry, dict) and entry.get("level") is not None)
    mode = f"populated from {rel_to_kit(Path(args.respostas))} ({answered} levels)" \
        if args.respostas else "empty template"
    print(f"Workbook generated -> {rel_to_kit(out_path)}")
    print(f"  Sheets: Leia-me | Respostas ({total_questions} questions) | Cálculo")
    print(f"  Mode: {mode}")
    print("  Weights: from framework.json (single source of truth)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
