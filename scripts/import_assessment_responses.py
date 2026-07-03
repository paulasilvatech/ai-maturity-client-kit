#!/usr/bin/env python3
"""Import AI Maturity Assessment responses from a Microsoft Forms Excel export.

Reads: respostas-forms.xlsx (or the path passed via --input)
Writes:
  - respostas.json (with automatic backup respostas.json.backup-<timestamp>)
  - saida/import-log-<DATE>.md (import log)

Rules (deterministic, mirrors /importar-respostas-excel skill):
  - Columns are matched by header regex P[1-3]-C<n>-Q<n> (column-order agnostic).
  - Evidence columns are matched by "(<qid>)" anywhere in the header, so
    "Evidência (P1-C1-Q1)" / "Evidence (P1-C1-Q1)" both work.
  - Multi-respondent aggregation: simple mean per question, NO rounding
    (2.5 is a legal level for scoring); evidences concatenated with
    "[respondent]: " prefix.
  - Empty / unmappable cells become null; values are never invented.
  - metadata.language: --language wins; otherwise the language already in
    respostas.json is preserved; "pt-br" is used only for brand-new files.

Usage:
    python3 import_assessment_responses.py
    python3 import_assessment_responses.py --input caminho/para/respostas-forms.xlsx
    python3 import_assessment_responses.py --language en
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import shutil
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
KIT = SCRIPT_DIR.parent  # kit root

QID_PATTERN = re.compile(r"^\s*(P[1-3]-C\d+-Q\d+)\b")
EVIDENCE_PATTERN = re.compile(r"\((P[1-3]-C\d+-Q\d+)\)")
NAME_HEADERS = {"name", "nome", "nombre"}
EMAIL_HEADERS = {"email", "e-mail", "correo", "correo electronico", "correo electrónico"}
VALID_LANGUAGES = ("pt-br", "en", "es")
MIN_QUESTIONS = 100  # below this the file is probably not a Forms export of the assessment


def fail(msg: str) -> "sys.NoReturn":
    print(f"❌ {msg}")
    sys.exit(1)


def load_workbook_safe(path: Path):
    """Open the xlsx, translating library errors into one-line messages."""
    try:
        import openpyxl
    except ImportError:
        fail("openpyxl não está instalado. Rode: pip install openpyxl")
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # zipfile.BadZipFile, InvalidFileException, etc.
        fail(f"Não consegui abrir '{path}' como .xlsx ({exc.__class__.__name__}). "
             f"Verifique se o arquivo é um export válido do Microsoft Forms.")


def load_json_safe(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"JSON inválido em {path}: {exc}. Corrija o arquivo antes de importar.")


def parse_level(cell_value) -> "float | None":
    """Forms exports the FULL option text, e.g. 'L3 - Gerenciado - >75% ...'."""
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    if not s:
        return None
    m = re.match(r"^L([0-4])\b", s)
    if m:
        return float(m.group(1))
    if s.upper().startswith("NA") or s.lower() in ("não sei", "nao sei", "n/a", "na"):
        return None  # explicit not-applicable
    return None  # unknown value; caller logs a warning


def framework_qids() -> list:
    fw = load_json_safe(KIT / "framework.json")
    qids = []
    for pillar in fw.get("pillars", []):
        for cap in pillar.get("capabilities", []):
            for q in cap.get("questions", []):
                qids.append(q["id"])
    if not qids:
        fail("framework.json não contém questões (pillars/capabilities/questions).")
    return qids


def map_columns(header_row):
    """Return (col_to_qid, col_to_evidence_qid, name_col, email_col), 1-based."""
    col_to_qid, col_to_ev = {}, {}
    name_col = email_col = None
    for idx, cell in enumerate(header_row, start=1):
        val = str(cell.value or "").strip()
        if not val:
            continue
        m = QID_PATTERN.match(val)
        if m:
            col_to_qid[idx] = m.group(1)
            continue
        em = EVIDENCE_PATTERN.search(val)
        if em:
            col_to_ev[idx] = em.group(1)
            continue
        low = val.lower()
        if low in NAME_HEADERS and name_col is None:
            name_col = idx
        elif low in EMAIL_HEADERS and email_col is None:
            email_col = idx
    return col_to_qid, col_to_ev, name_col, email_col


def collect_respondents(ws, col_to_qid, col_to_ev, name_col, email_col):
    """One respondent per data row; unanswered cells are simply skipped."""
    ev_by_qid = {qid: col for col, qid in col_to_ev.items()}
    respondents, warnings = [], []
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        def cell_value(col):
            return row[col - 1].value if col and col <= len(row) else None

        name = str(cell_value(name_col) or "").strip()
        email = str(cell_value(email_col) or "").strip()
        responses = {}
        for col, qid in col_to_qid.items():
            raw = cell_value(col)
            level = parse_level(raw)
            if level is None and raw is not None and str(raw).strip():
                warnings.append(
                    f"linha {row_idx}: valor não reconhecido em {qid} → {str(raw).strip()!r}, tratado como null")
            ev_col = ev_by_qid.get(qid)
            evidence = str(cell_value(ev_col) or "").strip() if ev_col else ""
            if level is not None or evidence:
                responses[qid] = {"level": level, "evidence": evidence}
        if not responses and not name and not email:
            continue  # fully empty row
        if not name:
            name = email or f"Respondente {len(respondents) + 1}"
        respondents.append({"name": name, "email": email, "responses": responses})
    return respondents, warnings


def aggregate(qids, respondents):
    """Mean of levels per question (no rounding); evidences concatenated."""
    agg = {}
    for qid in qids:
        levels, evidences = [], []
        for r in respondents:
            body = r["responses"].get(qid)
            if not body:
                continue
            if body["level"] is not None:
                levels.append(body["level"])
            if body["evidence"]:
                evidences.append(f"[{r['name']}]: {body['evidence']}")
        agg[qid] = {
            "level": (sum(levels) / len(levels)) if levels else None,
            "evidence": "\n".join(evidences),
            "n_respondents": len(levels),
        }
    return agg


def resolve_language(flag_value, existing_meta, file_is_new):
    if flag_value:
        return flag_value.lower()
    existing = (existing_meta or {}).get("language")
    if not file_is_new and existing:
        return existing  # preserve verbatim (e.g. "pt-BR")
    return "pt-br"  # brand-new file only


def build_log(date, source_name, respondents, agg, qids, backup_name, warnings,
              unmatched_qids):
    total = len(qids)
    answered = sum(1 for q in qids if agg[q]["n_respondents"] > 0)
    lines = [
        f"# Import log - {date}",
        "",
        "## Resumo",
        f"- Arquivo importado: {source_name}",
        f"- Respondentes: {len(respondents)} ({', '.join(r['name'] for r in respondents)})",
        f"- Questões processadas: {total} / {total}",
        f"- Questões com pelo menos 1 resposta: {answered}",
        f"- Backup do respostas.json anterior: {backup_name or '(arquivo novo, sem backup)'}",
        "",
        "## Cobertura por respondente",
        "| Respondente | Email | Respondidas | Evidências |",
        "|---|---|---|---|",
    ]
    for r in respondents:
        n_ans = sum(1 for b in r["responses"].values() if b["level"] is not None)
        n_ev = sum(1 for b in r["responses"].values() if b["evidence"])
        lines.append(f"| {r['name']} | {r['email'] or '(vazio)'} | {n_ans} / {total} | {n_ev} |")
    lines += ["", "## Alertas"]
    alerts = list(warnings)
    for qid in unmatched_qids:
        alerts.append(f"questão {qid} presente no Excel mas ausente do respostas.json, ignorada")
    unanswered = [q for q in qids if agg[q]["n_respondents"] == 0]
    if unanswered:
        alerts.append(
            f"{len(unanswered)} questões sem resposta de nenhum respondente, permanecem null "
            f"(ex.: {', '.join(unanswered[:5])}{'...' if len(unanswered) > 5 else ''})")
    if alerts:
        lines += [f"- {a}" for a in alerts]
    else:
        lines.append("- (nenhum): todos os valores foram parseados com sucesso")
    lines += ["", "## Próximo passo",
              "Rode `/pipeline-completo` para calcular scores e gerar relatório.", ""]
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(
        description="Importa respostas do assessment (export .xlsx do Microsoft Forms) "
                    "para respostas.json, agregando múltiplos respondentes por média.")
    ap.add_argument("--input", default=str(KIT / "respostas-forms.xlsx"),
                    help="Caminho do .xlsx exportado do Forms (default: respostas-forms.xlsx na raiz)")
    ap.add_argument("--output", default=str(KIT / "respostas.json"),
                    help="Caminho do respostas.json de saída (default: respostas.json na raiz)")
    ap.add_argument("--language", choices=VALID_LANGUAGES, type=str.lower, default=None,
                    help="Idioma dos relatórios (metadata.language). Default: preserva o idioma "
                         "já configurado no respostas.json existente; pt-br apenas para arquivo novo.")
    ap.add_argument("--log-dir", default=str(KIT / "saida"),
                    help="Diretório do import-log (default: saida/)")
    args = ap.parse_args()

    xlsx = Path(args.input)
    out_path = Path(args.output)
    log_dir = Path(args.log_dir)

    if not xlsx.exists():
        fail(f"Não encontrei o arquivo '{xlsx}'. Verifique o caminho ou rode com "
             f"--input <caminho-do-xlsx>.")

    wb = load_workbook_safe(xlsx)
    ws = wb.active
    col_to_qid, col_to_ev, name_col, email_col = map_columns(ws[1])

    if not col_to_qid:
        fail("Nenhum header no padrão P[1-3]-C<n>-Q<n> encontrado na linha 1. "
             "Confirme que o arquivo é um export do Forms do assessment "
             "(veja coleta/perguntas-para-forms.md).")
    if len(col_to_qid) < MIN_QUESTIONS:
        fail(f"Apenas {len(col_to_qid)} questões detectadas (esperado ~158). "
             f"O arquivo parece incompleto: verifique o formato dos headers.")

    qids = framework_qids()
    respondents, warnings = collect_respondents(ws, col_to_qid, col_to_ev, name_col, email_col)
    if not respondents:
        fail("Nenhuma linha de respondente encontrada (linhas 2+ estão vazias).")

    agg = aggregate(qids, respondents)
    unmatched = sorted(set(col_to_qid.values()) - set(qids))

    # Template: existing respostas.json > respostas.json.example > minimal skeleton
    file_is_new = not out_path.exists()
    if out_path.exists():
        template = load_json_safe(out_path)
    elif (KIT / "respostas.json.example").exists():
        template = load_json_safe(KIT / "respostas.json.example")
        template["target_overrides"] = {}
    else:
        template = {"metadata": {}, "target_overrides": {},
                    "responses": {q: {"level": None, "evidence": ""} for q in qids}}

    language = resolve_language(args.language, template.get("metadata"), file_is_new)

    # Backup before overwriting
    backup_name = None
    if out_path.exists():
        ts = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%S")
        backup_name = f"{out_path.name}.backup-{ts}"
        shutil.copy(out_path, out_path.parent / backup_name)

    today = datetime.date.today().isoformat()
    old_meta = template.get("metadata", {}) or {}
    single = respondents[0] if len(respondents) == 1 else None
    template["metadata"] = {
        "respondent_name": single["name"] if single
        else f"Agregado de {len(respondents)} respondentes",
        "respondent_email": single["email"] if single else "-",
        "respondent_role": "" if single else "Multi-respondente",
        "audience": ["all"],
        "organization": old_meta.get("organization", ""),
        "assessment_date": today,
        "language": language,
        "source": "microsoft-forms-import",
        "source_file": xlsx.name,
        "respondents": [{"name": r["name"], "email": r["email"]} for r in respondents],
    }
    for qid, body in agg.items():
        entry = template.setdefault("responses", {}).setdefault(qid, {"level": None, "evidence": ""})
        entry["level"] = body["level"]
        entry["evidence"] = body["evidence"]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(template, ensure_ascii=False, indent=2) + "\n",
                        encoding="utf-8")

    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"import-log-{today}.md"
    log_path.write_text(
        build_log(today, xlsx.name, respondents, agg, qids, backup_name, warnings, unmatched),
        encoding="utf-8")

    answered = sum(1 for q in qids if agg[q]["n_respondents"] > 0)
    n_ev = sum(1 for q in qids if agg[q]["evidence"])
    print(f"✓ Importação concluída → {out_path}")
    if backup_name:
        print(f"✓ Backup: {backup_name}")
    print(f"✓ Log: {log_path}")
    print(f"   • {len(respondents)} respondente(s): {', '.join(r['name'] for r in respondents)}")
    print(f"   • {answered} / {len(qids)} questões com pelo menos 1 resposta")
    print(f"   • {n_ev} evidências capturadas")
    print(f"   • Idioma dos relatórios (metadata.language): {language}")
    if warnings:
        print(f"⚠ {len(warnings)} alerta(s): valores não reconhecidos tratados como null (ver log)")
    print("Próximo: /pipeline-completo")
    return 0


if __name__ == "__main__":
    sys.exit(main())
