#!/usr/bin/env python3
"""Import Learning & Growth Survey responses from a Microsoft Forms Excel export.

Reads: respostas-survey-learning.xlsx (or the path passed via --input)
Writes:
  - survey-learning/respostas-learning.json
    (schema of survey-learning/respostas-mock-learning.json)
  - saida/import-learning-log-<DATE>.md (import log)

Rules (deterministic, mirrors /importar-survey-learning skill):
  - Columns matched by header regex L<n>-Q<n> (column-order agnostic).
  - The survey is IDENTIFIED by design: name and email are preserved per
    respondent (needed to invite people to workshops). The output metadata
    carries "contains_pii": true; handle and share the file accordingly.
  - Name/email come from the Forms Name/Email columns, falling back to the
    L1-Q1 (name) / L1-Q2 (email) answers. Rows with neither are skipped.
  - Question types (choice / multi / text / text-short) come from the bank
    survey-learning/perguntas-para-forms-learning.md (32 questions, 7 sections).

Usage:
    python3 import_learning_survey.py
    python3 import_learning_survey.py --input caminho/para/respostas-survey-learning.xlsx
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
KIT = SCRIPT_DIR.parent.parent  # kit root

QID_PATTERN = re.compile(r"^\s*(L[1-7]-Q\d+)\b")
BANK_HEADING = re.compile(r"###\s+Pergunta\s+`(L\d+-Q\d+)`\s+[—–-]+\s+_(.+?)_")
NAME_HEADERS = {"name", "nome", "nombre"}
EMAIL_HEADERS = {"email", "e-mail", "correo", "correo electronico", "correo electrónico"}


def fail(msg: str) -> "sys.NoReturn":
    print(f"❌ {msg}")
    sys.exit(1)


def load_workbook_safe(path: Path):
    try:
        import openpyxl
    except ImportError:
        fail("openpyxl não está instalado. Rode: pip install openpyxl")
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        fail(f"Não consegui abrir '{path}' como .xlsx ({exc.__class__.__name__}). "
             f"Verifique se o arquivo é um export válido do Microsoft Forms.")


def question_types() -> dict:
    """qid -> choice | multi | text | text-short, from the canonical question bank."""
    bank = KIT / "survey-learning" / "perguntas-para-forms-learning.md"
    if not bank.exists():
        fail(f"Banco de questões não encontrado: {bank}")
    mapping = {}
    for qid, kind in BANK_HEADING.findall(bank.read_text(encoding="utf-8")):
        low = kind.lower()
        if "multiple" in low:
            mapping[qid] = "multi"
        elif "single" in low:
            mapping[qid] = "choice"
        elif low.startswith("short text"):
            mapping[qid] = "text-short"
        else:
            mapping[qid] = "text"
    if not mapping:
        fail(f"Não consegui extrair os tipos de questão de {bank}.")
    return mapping


def map_columns(header_row):
    col_to_qid = {}
    name_col = email_col = None
    for idx, cell in enumerate(header_row, start=1):
        val = str(cell.value or "").strip()
        if not val:
            continue
        m = QID_PATTERN.match(val)
        if m:
            col_to_qid[idx] = m.group(1)
            continue
        low = val.lower()
        if low in NAME_HEADERS and name_col is None:
            name_col = idx
        elif low in EMAIL_HEADERS and email_col is None:
            email_col = idx
    return col_to_qid, name_col, email_col


def main():
    ap = argparse.ArgumentParser(
        description="Importa o Learning & Growth Survey (export .xlsx do Microsoft Forms, "
                    "IDENTIFICADO com nome+email) para survey-learning/respostas-learning.json.")
    ap.add_argument("--input", default=str(KIT / "respostas-survey-learning.xlsx"),
                    help="Caminho do .xlsx (default: respostas-survey-learning.xlsx na raiz)")
    ap.add_argument("--output", default=str(KIT / "survey-learning" / "respostas-learning.json"),
                    help="JSON de saída (default: survey-learning/respostas-learning.json)")
    ap.add_argument("--log-dir", default=str(KIT / "saida"),
                    help="Diretório do import-log (default: saida/)")
    args = ap.parse_args()

    xlsx = Path(args.input)
    out_path = Path(args.output)
    log_dir = Path(args.log_dir)

    if not xlsx.exists():
        fail(f"Não encontrei o arquivo '{xlsx}'. Veja survey-learning/INSTRUCOES-FORMS-LEARNING.md "
             f"ou rode com --input <caminho-do-xlsx>.")

    types = question_types()
    expected = len(types)  # 32 questions in the canonical bank

    wb = load_workbook_safe(xlsx)
    ws = wb.active
    col_to_qid, name_col, email_col = map_columns(ws[1])

    if not col_to_qid:
        fail("Nenhum header no padrão L<n>-Q<n> encontrado na linha 1. "
             "Confirme que o arquivo é o export do Learning & Growth Survey.")

    warnings = []
    if len(col_to_qid) != expected:
        warnings.append(f"{len(col_to_qid)} questões detectadas no Excel (esperado {expected} "
                        f"do banco canônico): verifique o formato dos headers")
    unknown = sorted(set(col_to_qid.values()) - set(types))
    if unknown:
        warnings.append(f"questões fora do banco canônico: {', '.join(unknown)}, "
                        f"importadas com type 'text'")

    qid_cols = {qid: col for col, qid in col_to_qid.items()}
    respondents = []
    skipped_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]

        def cell_value(col):
            return row[col - 1].value if col and col <= len(row) else None

        responses = {}
        for col, qid in col_to_qid.items():
            raw = cell_value(col)
            if raw is None or str(raw).strip() == "":
                continue
            responses[qid] = {"type": types.get(qid, "text"), "value": str(raw).strip()}

        name = str(cell_value(name_col) or "").strip()
        email = str(cell_value(email_col) or "").strip()
        # The L1-Q1/L1-Q2 answers are the better source when Forms columns are empty.
        if not name:
            name = responses.get("L1-Q1", {}).get("value", "")
        if not email:
            email = responses.get("L1-Q2", {}).get("value", "")

        if not responses:
            continue  # fully empty row
        if not name and not email:
            skipped_rows.append(row_idx)
            continue  # identified survey: a row without any identity is incomplete
        respondents.append({
            "respondent_id": len(respondents) + 1,
            "name": name,
            "email": email,
            "responses": responses,
        })

    if not respondents:
        fail("Nenhum respondente identificado encontrado (linhas vazias ou sem nome/email).")
    if skipped_rows:
        warnings.append(f"linhas puladas por falta de nome E email (colunas Name/Email e "
                        f"L1-Q1/L1-Q2 vazias): {', '.join(str(r) for r in skipped_rows)}")

    output = {
        "metadata": {
            "source": xlsx.name,
            "imported_at": datetime.datetime.now(datetime.timezone.utc)
                           .isoformat(timespec="seconds").replace("+00:00", "Z"),
            "total_respondents": len(respondents),
            "total_questions": expected,
            "total_questions_detected": len(col_to_qid),
            "anonymous": False,
            "scope": "individual_identified (name+email per respondent)",
            "contains_pii": True,
        },
        "respondents": respondents,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n",
                        encoding="utf-8")

    champions = sum(
        1 for r in respondents
        if r["responses"].get("L6-Q1", {}).get("value", "").lower().startswith("sim"))

    today = datetime.date.today().isoformat()
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"import-learning-log-{today}.md"
    total_cells = len(col_to_qid) * len(respondents)
    got_cells = sum(len(r["responses"]) for r in respondents)
    coverage = round(100 * got_cells / total_cells) if total_cells else 0
    log_lines = [
        f"# Import log - Learning & Growth Survey ({today})",
        "",
        "> Atenção: o arquivo gerado contém dados pessoais (nome + email).",
        "> Distribuição restrita, não compartilhar publicamente.",
        "",
        "## Resumo",
        f"- Arquivo: {xlsx.name}",
        f"- Respondentes: {len(respondents)} (IDENTIFICADOS, nome+email preservados)",
        f"- Questões detectadas: {len(col_to_qid)} / {expected}",
        f"- Cobertura de respostas: {coverage}%",
        f"- Candidatos a Champion (L6-Q1 = Sim): {champions}",
        f"- Saída: {out_path.relative_to(KIT) if out_path.is_relative_to(KIT) else out_path}",
        "",
        "## Alertas",
        *([f"- {w}" for w in warnings] or ["- (nenhum)"]),
        "",
        "## Próximo",
        "Rode `/plano-capacitacao` para gerar o plano priorizado em saida/.",
        "",
    ]
    log_path.write_text("\n".join(log_lines), encoding="utf-8")

    print(f"✓ Survey de Learning importado → {out_path}")
    print(f"   {len(respondents)} respondentes IDENTIFICADOS, {len(col_to_qid)} questões processadas")
    print(f"✓ Log: {log_path}")
    print(f"   Cobertura: {coverage}%")
    print(f"   Champions candidates (L6-Q1 = Sim): {champions}")
    print("⚠ O JSON contém dados pessoais (contains_pii: true), distribuição restrita")
    if warnings:
        print(f"⚠ {len(warnings)} alerta(s), ver log")
    print("Próximo: /plano-capacitacao")
    return 0


if __name__ == "__main__":
    sys.exit(main())
